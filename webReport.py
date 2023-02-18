import streamlit as st
import openpyxl as xl
import matplotlib.pyplot as plt
from string import capwords

# studentDataWB = xl.load_workbook(r'studentCost.xlsx')
# invoiceDataWB = xl.load_workbook(r'InvoiceList_Jan2023.xlsx')

studentDataWB = xl.load_workbook(r'studentData.xlsx')
invoiceDataWB = xl.load_workbook(r'InvoiceList_Feb2023.xlsx')

studentData = {}
invoiceData = {}
overDueData = {}


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


for sheet in studentDataWB.sheetnames:
    stName = []
    # print(studentDataWB)
    studentDataWS = studentDataWB[sheet]
    maxRow = get_maximum_rows(sheet_object=studentDataWS)
    for row in range(2,maxRow+1):
        stName.append(capwords(studentDataWS['B' + str(row)].value))
    studentData[sheet] = stName

for sheetInvoice in invoiceDataWB.sheetnames:
    stNameInvoice = []
    # print(invoiceDataWB)
    invoiceDataWS = invoiceDataWB[sheetInvoice]
    maxRow = get_maximum_rows(sheet_object=invoiceDataWS)
    for row in range(2,maxRow+1):
        stNameInvoice.append(capwords(invoiceDataWS['A' + str(row)].value))
    # print(stNameInvoice)
    invoiceData[sheetInvoice] = stNameInvoice

for className in studentData:
    paidStudent = invoiceData.get(className)
    totalStudent = studentData.get(className)
    unpaidStudent = []
    # print(paidStudent)

    for student in totalStudent:
        if student not in paidStudent:
            unpaidStudent.append(student)

    overDueData[className] = unpaidStudent


for className in studentData:
    invoiceReport = [len(overDueData.get(className)), len(invoiceData.get(className))]
    invoiceReportLabel = ['\n'.join(overDueData.get(className)), '\n'.join(invoiceData.get(className))]
    invoiceReportLegendLabel = ['Unpaid', 'Paid']
    # plt.pie(invoiceReport, labels=invoiceReportLabel, startangle=270, colors=['r', 'g'])
    # plt.legend(labels=invoiceReportLegendLabel, title=className, loc=1)
    # plt.savefig(className + '.png')
    # plt.show()

    fig1, ax1 = plt.subplots()
    ax1.pie(invoiceReport, labels=invoiceReportLabel, autopct='%1.1f%%', startangle=270, colors=['r','g'])
    ax1.legend(labels=invoiceReportLegendLabel, title=className, loc=1)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    st.pyplot(fig1)



# Pie chart, where the slices will be ordered and plotted counter-clockwise:
# labels = 'Frogs', 'Hogs', 'Dogs', 'Logs'
# sizes = [15, 30, 45, 10]
# explode = (0, 0.1, 0, 0)  # only "explode" the 2nd slice (i.e. 'Hogs')
